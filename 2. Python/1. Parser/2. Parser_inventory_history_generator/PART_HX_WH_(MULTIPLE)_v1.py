from openpyxl import Workbook, load_workbook
import os
from datetime import datetime, date

############################################################
#
#               Part Flow Check by Stan
#               V 1.0_2022_02_26 (Multi Warehouse Version)
#
############################################################

row_num = { 0:'A', 1:'B', 2:'C', 3:'D', 4:'E', 5:'F', 6:'G', 7:'H', 8:'I', 9:'J', 10:'K', 11:'L', 12:'M', 13:'N', 14:'O', 15:'P', 16:'Q', 17:'R', 18:'S', 19:'T', 20:'U', 21:'V', 22:'W', 23:'X', 24:'Y', 25:'Z' }
NO = 0

def get_file_list():
    path = os.getcwd()
    print(path)

    mylist = []
    for path, dir, files in os.walk(path):
        for filename in files:
            if 'XLSX' in filename:
                mylist.append(filename.split('.')[0])
    return mylist

def get_data(wb_name, warehouse, part_no):
    date_no = wb_name                                                           # store the date value
    wb_name += ".XLSX"
    in_workbook = load_workbook(wb_name)
    in_worksheet = in_workbook.active
    print(f'\nFILE NAME: {wb_name} -------------------------------------------------> READ FILE CORRECT')
    print(f'FILE TYPE: {type(in_worksheet)} ----------> READ FILE CORRECT')
    print(f'<----- ↓ FILE PROCESS RESULT ↓ ----->')

    global NO
    idx = 4                                                                     # default idx 4 
    data_list = []                                                              # empty list to store data for later data delivery
    tmp_list = []
    
    # <--- section to matching and locate the target warehouse col --->
    # locate part no col
    i = 0
    part_head_idx = row_num[i] + str(idx)
    while ( None != in_worksheet[part_head_idx].value ):

        part_head_idx = row_num[i] + str(idx)

        if ( "品號" == in_worksheet[part_head_idx].value ):
            print(f'COL INDEX FOUND FOR 品號 ------------------------------> {part_head_idx}')
            part_col_idx = i                                               # store the col value of part no col
            break

        if ( i > 25 ):
            print('ERROR! 品號 NOT FOUND!')
            part_col_idx = 0
            break

        i += 1

    # locate part no row
    part_row_idx = idx
    part_idx = row_num[part_col_idx] + str(part_row_idx)

    while ( None != in_worksheet[part_idx].value ):
        part_idx = row_num[part_col_idx] + str(part_row_idx)
        if ( in_worksheet[part_idx].value == part_no ):
            print(f'COL INDEX FOUND FOR PART NO:{part_no} ---------------> {part_idx}')
            tmp_list += [in_worksheet[part_idx].value]
            break
        if ( None == in_worksheet[part_idx].value ):
            tmp_list += ['None']
        part_row_idx += 1

    # Locate part qty in each warehouse
    for wh in warehouse:
        j = 0
        warehouse_idx = row_num[j] + str(idx)
        while ( None != in_worksheet[warehouse_idx].value ):
            warehouse_idx = row_num[j] + str(idx)
            if ( in_worksheet[warehouse_idx].value == wh ):
                print(f'COL INDEX FOUND FOR WAREHOUSE: {wh}------------------------------> {warehouse_idx}')
                part_qty_idx = row_num[j] + str(part_row_idx)                            #  location of specific part_qty in designate wh
                print(f'COL INDEX FOUND FOR PART: {part_no} IN {wh}-------------------> {part_qty_idx}')
                part_qty = in_worksheet[part_qty_idx].value
                tmp_list += [part_qty]
                break
            j += 1

    # store the parsed data to a list
    NO += 1
    data_list.append(NO)
    data_list.append(date_no)
    data_list.extend(tmp_list)

    print(f'DATA PARSED: {data_list}')

    return data_list

def get_file(warehouse, part_no):
    
    #create a new workbook and worksheet for data storage
    out_workbook = Workbook()
    out_workwheet = out_workbook.active
    out_workwheet.title = 'HX_'+part_no
    
    # receives the filelist in the folder
    file_list = get_file_list()
    print(f'File Lists: {file_list}\n<---------- READ FILE CORRECCT ---------->')
    header = [ 'No', 'Date', 'Part No']
    header.extend(warehouse)
    out_workwheet.append(header)
    
    try :
        # loop in to get data and append in worksheet
        for wb_name in file_list:
            tmp_list = get_data(wb_name, warehouse, part_no)
            out_workwheet.append(tmp_list)
    except :
        print("< --- Fail to Read Any File --- >")

    return out_workbook

def time_stamp():

    today = date.today()
    now = datetime.now().time()

    today = list(str(today))
    now = list(str(now))

    print(today)
    print(now)

    for i in range(len(now)):
        if ( '.' == now[i] or ':'== now[i] ):
            now[i] = '_'

    for j in range(len(today)):
        if ( '-' == today[j]):
            today[j] ='_'

    date_stamp = ''.join(today)
    time_stamp = ''.join(now)
    stamp = date_stamp + '_(' + time_stamp +')'

    return stamp

def create_folder():
    
    path = os.getcwd()
    os.chdir(path)
    stamp = time_stamp()
    new_folder = "HX_RESULT_"+stamp
    os.makedirs(new_folder)

    new_folder = "\HX_RESULT_"+stamp
    return new_folder


def portal(warehouse, part_no, folder):
    # call get_list to return a combined list
    global NO
    NO = 0

    path = os.getcwd()
    out_workbook = get_file(warehouse, part_no)
    filename = path + folder + "\HX_"+part_no+'.xlsx'
    out_workbook.save(filename)

def main():
    wh_list = ['內湖倉','信義倉','待整新倉','配貨倉','國際倉']
    part_list = ['HCM009APWW2F', 'HCM006APWW2F', 'HCM005APWW2F', 'HCM004APWW2F', 'HBC011ZZC206', 'HBC011ENC206', 'HBC011ARC206', 'HBC008ZZC201', 'HBC008JGC106', 'HBC007JGC106', 'HBC006JGC106', 'HBC005ZZC106']
    folder = create_folder()
    for part_no in part_list:
        portal(wh_list, part_no, folder)

if (__name__ == '__main__'):
    main()