import os
from datetime import datetime, date
from openpyxl import Workbook, load_workbook

############################################################
#
#               Inventory Generator by Stan
#               V 1.0
#               DATE : 2022_06_28
#
############################################################

global_time_stamp = ""
total_inventory:int = 0
single_item_inventory_sum:int = 0

#################### General Functions ####################

# GET A COMPLETE TIME STAMP: 2022_06_28_(14_42_20_144672)
def get_time_stamp():

    today = date.today()
    now = datetime.now().time()

    today = list(str(today))
    now = list(str(now))

    for i in range(len(now)):
        if ( '.' == now[i] or ':'== now[i] ):
            now[i] = '_'

    for j in range(len(today)):
        if ( '-' == today[j]):
            today[j] ='_'

    date_stamp = ''.join(today)
    time_stamp = ''.join(now)
    
    stamp = date_stamp + '_(' + time_stamp[0:8] +')'

    return stamp

# GET EXCEL COLUMN INDEX SERIES STRING
def get_excel_series_string(num:int):
    num_char = ""
    A = 65
    while ( num >= 0 ):    
        rem = num % 26
        num_char += chr( A + rem )
        num = (num//26) - 1
   
    num_char = num_char[::-1]
    return num_char

def get_warehouse_list(wb_name):

    warehouse_list = []

    wh_criterion  = '庫別名稱'
    start_of_header = '品號'
    wb_name += ".XLSX"
    in_workbook = load_workbook(wb_name)
    in_worksheet = in_workbook.active

    worksheet_in_use_row_count = in_worksheet.max_row

    log = f' WORKSHEET ROWS ---------> {worksheet_in_use_row_count}'
    get_log(log)
    
    excel_series_num = 0
    row_idx = 1
    
    # LOOP TO LOCATE HEADER ROW_IDX: A4
    while (True):
        cell_idx = get_excel_series_string(excel_series_num) + str(row_idx)
        if ( start_of_header == in_worksheet[cell_idx].value ):
            break
        row_idx += 1

    # LOOP TO LOCATE WHERE THE COLOUMN OF WAREHOUSE NAME IS [COL_ID: EXCEL SERIES NUM]
    while (True):
        cell_idx = get_excel_series_string(excel_series_num) + str(row_idx)
        if ( wh_criterion == in_worksheet[cell_idx].value ):
            break
        excel_series_num += 1

    # LOOP TO LOCATE EACH WAREHOUSE NAME AND STORE TO A NON-REPEAT LIST
    row_idx += 1
    for i in range(worksheet_in_use_row_count):
        cell_idx = get_excel_series_string(excel_series_num) + str(row_idx)
        warehouse_name = str(in_worksheet[cell_idx].value)
        
        if ( '倉' in warehouse_name ):
            if ( warehouse_name not in warehouse_list):
                warehouse_list.append(warehouse_name)
        row_idx += 1

    text_sum = '品項總計'
    warehouse_list.append(text_sum)
    log = f' WAREHOUSE FOUND --------> {str(warehouse_list)}'
    get_log(log)

    return warehouse_list

#################### FILE PROCESS ####################
def get_file_list():
    path = os.getcwd()
    file_list = []
    for path, dir, files in os.walk(path):
        for filename in files:
            if 'XLSX' in filename:
                file_list.append(filename.split('.')[0])
    return file_list

def get_folder_name_and_create_folder():
    path = os.getcwd()
    os.chdir(path)
    global global_time_stamp
    global_time_stamp = get_time_stamp()
    new_folder = "INV_"+ global_time_stamp
    os.makedirs(new_folder)
    new_folder = "\INV_"+ global_time_stamp
    return new_folder  # \INV_2022_06_28_(15_24_20_293669)

def get_file_name(folder_name):
    path = os.getcwd()
    filename = path + folder_name + folder_name+'.xlsx'
    return filename

def get_log(string:str):
    global global_time_stamp
    path = os.getcwd()
    log_file_name = path +  "\INV_"+ global_time_stamp + '\Log_INV' + global_time_stamp + '.txt'
    with open(log_file_name, 'a') as file:
        file.write(string + '\n')

def load_worksheet(wb_name):
    wb_name += ".XLSX"
    in_workbook = load_workbook(wb_name)
    in_worksheet = in_workbook.active
    return in_worksheet

def get_workbook_row_count(wb_name):
    wb_name += ".XLSX"
    in_workbook = load_workbook(wb_name)
    in_worksheet = in_workbook.active
    worksheet_in_use_row_count = in_worksheet.max_row
    return worksheet_in_use_row_count

def get_workbook_col_count(wb_name):
    wb_name += ".XLSX"
    in_workbook = load_workbook(wb_name)
    in_worksheet = in_workbook.active
    worksheet_in_use_col_count = in_worksheet.max_column
    return worksheet_in_use_col_count

#################### CORE PROCESS ####################

# GET WAREHOUSE SUME
def get_warehouse_sum(sum_list:list, worksheet)->list:

    valid_col_size = worksheet.max_column
    valid_row_size = worksheet.max_row

    log = f' TOTAL ROW X COL --------> {[valid_row_size]} x {[valid_col_size]}'
    get_log(log)
    
    sum_list[0] = None
    sum_list[1] = '總計'
    for col_idx in range (2, valid_col_size):
        tmp_sum = 0
        for row_idx in range(2, valid_row_size+1):
            cell_idx = get_excel_series_string(col_idx) + str(row_idx)
            inventory = worksheet[cell_idx].value
            tmp_sum += inventory
            # print(f'{[cell_idx]} ---> {type(worksheet[cell_idx].value)}')
        sum_list[col_idx] = tmp_sum

    return sum_list


# DETAILED INVENTORY PROCESS WHEN VALID PART NO FOUND
def get_detailed_part_inv(data_list, col_idx, row_idx, worksheet, warehouse_list, worksheet_col_count):

    refined_data_list = []
    for i in range(len(warehouse_list)):
        refined_data_list.append(0)
    
    item = '0'

    # SEARCH FROM PASSED IN ROW_IDX TO FIND PART NO AND ITS INVENTORY IN WAREHOUSE
    while ( True ):

    ##### SEARCH FROM LEFTMOST TO RIGHTMOST TO FIND WAREHOUSE
        for col_idx in range(worksheet_col_count):

            cell_idx = get_excel_series_string(col_idx) + str(row_idx)
            item = str(worksheet[cell_idx].value)
            
    ########## WHEN WAREHOUSE FOUND FOR THE PART NO
            if ( '倉' in item ):

    ########## GET THE INVENTORY FROM TWO COLUMN AFTER THE CURRENT COLUMN
                col_idx += 2
                target_cell_idx = get_excel_series_string(col_idx) + str(row_idx)
                inventory = int(worksheet[target_cell_idx].value)
    
    ########### CHECK SUM OF THE INVENTORY
                global total_inventory 
                total_inventory += inventory

                global single_item_inventory_sum
                single_item_inventory_sum += inventory

    ########### LOCATE WHICH POSITION TO INSERT THE INVENTORY TO MATCH THE WAREHOUSE LIST HEADER
                target_insert_idx = warehouse_list.index(item)
                refined_data_list[target_insert_idx] = inventory

                log = f' INVENTORY --------------> {[ inventory ]} FOUND IN {[ item ]} AT CELL { [ target_cell_idx ] }'
                get_log(log)
                break

    ##### KEEP SEARCHING DOWN TO NEXT RWO TO FIND VALID INVENTORY IN WAREHOUSE FOR PART NO
        col_idx = 3
        row_idx += 1
        cell_idx = get_excel_series_string(col_idx) + str(row_idx)
        have_something = worksheet[cell_idx].value
        
        log = f' VERDICT TOKEN FOUND ----> {[ have_something ] } FOUND IN {[ cell_idx ]}'
        get_log(log)

        if ( not have_something ):
            log = f' VERDICT PROCESS END ----> {[ have_something ]}'
            get_log(log)
            break

    log = f' ITEM INVENTORY SUM -----> {[single_item_inventory_sum]}'
    get_log(log)

    refined_data_list[-1] = single_item_inventory_sum
    single_item_inventory_sum  = 0

    data_list.extend(refined_data_list)
    return data_list


# CHECK IF IT'S A VALID PART NO FOR FURTHER PROCESS
def get_part_info_list(col_idx, row_idx, worksheet, warehouse_list, worksheet_col_count):
    data_list = []

    #### START FROM LEFTMOST COL A to THE RIGHTMOST COL TO CHECK VALIDITY
    for col_idx in range(worksheet_col_count):
        cell_idx = get_excel_series_string(col_idx) + str(row_idx)
        warehouse_name = str(worksheet[cell_idx].value)
        
    ######## WHEN VALID INFO FOUND, PROCEED TO GET THE INVENTORY DATA LIST
        if ( '倉' in warehouse_name ):

            log = f' VALID DATA FOUND -------> {[ cell_idx ]} AS {[ warehouse_name ]}'
            get_log(log)

    ############ CREATE A PREPARED LIST WITH VALID PART NO AND PART DESCRIPTION FOR FURTHER INVENTORY LIST TO APPEND
            tmp_id = []
            for col_idx in range(2):
                cell_idx = get_excel_series_string(col_idx) + str(row_idx)
                item = worksheet[cell_idx].value
                data_list.append(item)
                tmp_id.append(cell_idx)

            log = f' DATA LIST PREPARED -----> {tmp_id} AS {data_list}'
            get_log(log)

            col_idx = 0
            
    ########### GET THE DETAILED INVENTORY LIST
            data_list = get_detailed_part_inv(data_list, col_idx, row_idx, worksheet, warehouse_list, worksheet_col_count)
            break    

    return data_list


def get_file(warehouse_list, sheet_name, file_name):
    
    #CREATE NEW WORKBOOK
    out_workbook = Workbook()
    out_workwheet = out_workbook.active
    out_workwheet.title = sheet_name

    log = f'\nFile Name: {file_name}\n<-------------------- READ FILE CORRECCT -------------------->'
    get_log(log)

    header = ['品號', '品名']
    header.extend(warehouse_list)
    out_workwheet.append(header)

    #################################### PROCESS DATA HERE AND APPEND TO THE WORKSHEET ########################################
    in_worksheet = load_worksheet(file_name)
    col_idx = 0;
    ret_list = []
    worksheet_row_count = get_workbook_row_count(file_name)
    worksheet_col_count = get_workbook_col_count(file_name)

    # SEARCH FROM TOP ROW to FINAL ROW
    for row_idx in range(1, worksheet_row_count + 1):
        cell_idx = get_excel_series_string(col_idx) + str(row_idx)

    ### FOUND SOME INFO TO FURTHER CHECK VALIDITY AND PROCESS
        if ( None != in_worksheet[cell_idx].value ):

            log = f' DATA FOUND -------------> {[ in_worksheet[cell_idx].value ]} IN CELL {[ cell_idx ]}'
            get_log(log)

    ####### GET THE PART INFO IF VALID PART NO FOUND
            ret_list = get_part_info_list(col_idx, row_idx, in_worksheet, warehouse_list, worksheet_col_count)

    ####### APPEND THE VALID DATA LIST TO THE WORKSHEET
            if (len(ret_list) != 0):

                log = f' PROCESSED DATA LIST ----> {ret_list}'
                get_log(log)
                log = '<---------------------------------------- [RESULT] VALID DATA PROCESS ----------------------------------------->\n'
                get_log(log)

                out_workwheet.append(ret_list)
            else:
                log = f' INVALID DATA RETURNED --> {ret_list}'
                get_log(log)
                log = '<---------------------------------------  [RESULT] INVALID DATA PROCESS --------------------------------------->\n'
                get_log(log)

    ##################################### COMPLETEION OF THE DATA PROCESS TASK ###################################################
    
    
    # GET SUM IN THE WORKSHEET
    log = f' SUM OF INVENTORY -------> {[ total_inventory ]}'
    get_log(log)

    sum_list = []
    for i in header:
        sum_list.append(0)

    # GET SUM LIST
    sum_list = get_warehouse_sum(sum_list, out_workwheet)

    sum_list[-1] = total_inventory
    out_workwheet.append(sum_list)

    return out_workbook

#################### Portal ####################
def main():

    print("+---------------------------------------+")
    print("|                                       |")
    print("|      INVENTORY TABLE GENERATOR        |")
    print("|            by Stan Liao               |")
    print("|            Version 1.0                |")
    print("|         Release: Jun 29 2022          |")
    print("|                                       |")
    print("+---------------------------------------+\n")

    print(" ** --> 庫存表產生中.....\n")

    file_names_list = get_file_list()
    for file_name in file_names_list:
        
        # CREATE A FOLDER AND CHECK WAREHOUSE MARKS
        folder_name = get_folder_name_and_create_folder()
        warehouse_list = get_warehouse_list(file_name)

        # GET THE PROCESSED WORKBOOK FROM GET FILE
        out_workbook = get_file(warehouse_list, folder_name[1:], file_name)

        log = '<----------------------------------------------- END OF LOG ----------------------------------------------->\n'
        get_log(log)

        # SAVE TO A NEW FOLDER IN DESIGNATED DIRECTORY
        file_path_dir_name = get_file_name(folder_name)
        out_workbook.save(file_path_dir_name)

        print(f' ** --> 庫存表 {folder_name[1:]} 製作完成.....\n')
        print(f' ** --> 詳情請查閱 Log_{folder_name[1:]} 處裡記錄檔 .....\n')
        print(f' ** --> 檔案位置：{file_path_dir_name}')

        os.system('pause')
        
if (__name__ == '__main__'):
    main()