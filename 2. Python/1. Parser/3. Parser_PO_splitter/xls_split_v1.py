from openpyxl import Workbook, load_workbook

###########################################################
#
#       xls_splitter_StudioA ver 1.0.0  by Stan
#
###########################################################

row_num = { 0:'A', 1:'B', 2:'C', 3:'D', 4:'E', 5:'F', 6:'G', 7:'H', 8:'I', 9:'J', 10:'K', 11:'L', 12:'M', 13:'N', 14:'O', 15:'P', 16:'Q', 17:'R', 18:'S', 19:'T', 20:'U', 21:'V', 22:'W', 23:'X', 24:'Y', 25:'Z' }

# <----- Load Workbook ----->
in_workbook = load_workbook('in.xlsx')
in_worksheet = in_workbook.active
print(f'\nInput File Confirmed: {type(in_worksheet)}\n')

# <----- Save to New Workbook ----->
def create_workbook(sheetname, in_worksheet, idx, header_row):
    out_workbook = Workbook()
    out_worksheet = out_workbook.active
    out_worksheet.title = sheetname
    out_worksheet.append(header_row)
    
    col_idx = 'C' + str(idx)
    # print(f'sheetname:{sheetname}, in_worksheet[{col_idx}]:{in_worksheet[col_idx].value}')

    # process to append row to the new worksheet if the value of worksheet[col_idx] is the same
    while( sheetname == in_worksheet[col_idx].value ):

        #create row to append
        row = []
        i=0
        row_idx = row_num[i] + str(idx)
        while ( None != in_worksheet[row_idx].value ):
            row_idx = row_num[i]+str(idx)
            row_val = in_worksheet[row_idx].value
            row.append(row_val)
            i += 1
            if ( i > 25 ):
                print('Row Size is too large. Must not be larger than 25 col.')
                break
        print(f'row of col_idx to append ---> {row}')

        # append row to the col
        idx += 1
        col_idx = 'C' + str(idx)
        out_worksheet.append(row)

    filename = sheetname + '.xlsx'
    out_workbook.save(filename)
    return idx

#<------ Header Row ------>
header_row = []
i=0
row_idx = row_num[i]+'1'
while ( None != in_worksheet[row_idx].value ):
    row_idx = row_num[i]+'1'
    row_val = in_worksheet[row_idx].value
    header_row.append(row_val)
    i += 1
    if ( i > 25 ):
        print('Row Size is too large. Must not be larger than 25 col.')
        break
print(f'Header Row ---> {header_row}')

# <----- Pre-set ----->
count = 0
idx = 2
col_idx = 'C' + str(idx)
current_col = in_worksheet[col_idx].value
print(f'The initial column index ---> {col_idx}')

# <----- Main Process ----->
table = []
while( None != in_worksheet[col_idx].value ):
    
    if ( in_worksheet[col_idx].value == current_col ):
        sheetname = in_worksheet[col_idx].value
        rc = create_workbook(sheetname, in_worksheet, idx, header_row)
        idx = rc

        col_idx = 'C' + str(idx)
        current_col = in_worksheet[col_idx].value
        print(f'\nrc--->{rc}, idx--->{idx}, col_idx--->{col_idx}, current_col--->{current_col}')

    col_idx = 'C' + str(idx)
    count+=1

print(f'Total File Count: {count}')